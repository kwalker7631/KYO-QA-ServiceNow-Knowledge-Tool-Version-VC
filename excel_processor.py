# excel_processor.py
# Author: Kenneth Walker
# Date: 2025-08-18 (Updated)
# Version: VA-1.7 (Improved Auto-Formatting)

import logging
import openpyxl
from util import excel_safe
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from datetime import datetime

# --- CONFIGURATION ---
# Expected column names (case-sensitive)
DESCRIPTION_COL = "Description"
SHORT_DESCRIPTION_COL = "Short description"
META_COL = "Meta"
AUTHOR_COL = "Author"
TOPIC_COL = "Topic"
PRODUCT_DESC_COL = "Product Description"
STATUS_COL = "Processing Status"
SYS_ID_COL = "Sys ID"

logger = logging.getLogger(__name__)

def process_excel_file(template_path: Path, processed_data: list, output_dir: Path) -> Path:
    """
    Clones and updates an Excel template with intelligent auto-formatting.
    It moves PDF names, updates or creates rows, and applies professional styling
    for readability, including column sizing, text wrapping, and alignment.

    Args:
        template_path (Path): Path to the user's ServiceNow Excel template.
        processed_data (list): List of dicts from main app, each for one PDF.
        output_dir (Path): Directory for the new Excel file.

    Returns:
        Path: Path to the populated Excel file.
    """
    if not processed_data:
        raise ValueError("No processed data provided; nothing to populate.")

    # 1. Create timestamped clone
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_filename = f"PROCESSED_{template_path.stem}_{timestamp}.xlsx"
    cloned_path = output_dir / output_filename

    try:
        workbook = openpyxl.load_workbook(template_path)
        excel_safe.save_workbook_safely(workbook, cloned_path)
    except Exception as e:
        raise IOError(f"Could not clone Excel template at '{template_path}'. Error: {e}")

    # 2. Load clone and find/add column indices
    workbook = openpyxl.load_workbook(cloned_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1] if cell.value]
    header_dict = {name: idx + 1 for idx, name in enumerate(headers)}

    # Ensure required columns exist and make headers bold
    required_cols = [DESCRIPTION_COL, SHORT_DESCRIPTION_COL, META_COL, AUTHOR_COL, TOPIC_COL, PRODUCT_DESC_COL, STATUS_COL]
    next_col = len(headers) + 1
    for col_name in required_cols:
        if col_name not in header_dict:
            cell = sheet.cell(row=1, column=next_col, value=col_name)
            cell.font = Font(bold=True)
            header_dict[col_name] = next_col
            next_col += 1
    
    # Make existing headers bold
    for col_idx in header_dict.values():
        sheet.cell(row=1, column=col_idx).font = Font(bold=True)


    # 3. Pre-process: Move any .pdf filenames from Short Description to Description
    if SHORT_DESCRIPTION_COL in header_dict and DESCRIPTION_COL in header_dict:
        short_desc_idx = header_dict[SHORT_DESCRIPTION_COL]
        desc_idx = header_dict[DESCRIPTION_COL]
        for row in range(2, sheet.max_row + 1):
            short_desc_cell = sheet.cell(row=row, column=short_desc_idx)
            if short_desc_cell.value and str(short_desc_cell.value).lower().endswith('.pdf'):
                if not sheet.cell(row=row, column=desc_idx).value:
                    sheet.cell(row=row, column=desc_idx).value = short_desc_cell.value
                    short_desc_cell.value = None

    # 4. Build a map of existing PDF descriptions to their row number
    description_to_row_map = {}
    if DESCRIPTION_COL in header_dict:
        desc_idx = header_dict[DESCRIPTION_COL]
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=desc_idx).value
            if cell_value:
                description_to_row_map[Path(str(cell_value)).stem] = row
    
    # 5. Update existing rows or append new ones
    data_map = {Path(item["filename"]).stem: item for item in processed_data}
    next_new_row = sheet.max_row + 1
    
    for pdf_stem, data in data_map.items():
        row_index = description_to_row_map.get(pdf_stem)
        if row_index is None:
            row_index = next_new_row
            next_new_row += 1

        def get_items(item_type):
            return ", ".join(sorted([item['text'] for item in data.get('found_items', []) 
                                     if item['type'].replace(" ", "_").lower() == item_type and not item['flagged']]))

        models = get_items("model")
        qa_numbers = get_items("qa_number")
        authors = get_items("author")
        topics = get_items("topic")

        sheet.cell(row=row_index, column=header_dict[DESCRIPTION_COL], value=data["filename"])
        sheet.cell(row=row_index, column=header_dict[SHORT_DESCRIPTION_COL], value=qa_numbers or pdf_stem)
        sheet.cell(row=row_index, column=header_dict[META_COL], value=models or "Not Found")
        sheet.cell(row=row_index, column=header_dict[AUTHOR_COL], value=authors)
        sheet.cell(row=row_index, column=header_dict[TOPIC_COL], value=topics)
        
        if PRODUCT_DESC_COL in header_dict:
            sheet.cell(row=row_index, column=header_dict[PRODUCT_DESC_COL], value="")
            
        sheet.cell(row=row_index, column=header_dict[STATUS_COL], value=data.get("status", "Fail"))

    # 6. Apply conditional formatting
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if STATUS_COL in header_dict:
        status_idx = header_dict[STATUS_COL]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            status_cell = sheet.cell(row=row[0].row, column=status_idx)
            if status_cell.value:
                fill_color = None
                if "Pass" in status_cell.value:
                    fill_color = pass_fill
                elif "Needs Review" in status_cell.value:
                    fill_color = review_fill
                elif "Fail" in status_cell.value:
                    fill_color = fail_fill
                
                if fill_color:
                    for cell in row:
                        cell.fill = fill_color
    
    # --- 7. RE-ENABLED: Apply improved auto-formatting ---
    # Set alignment and wrapping for all data cells
    align = Alignment(wrap_text=True, vertical='top', horizontal='left')
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = align

    # Auto-size columns based on content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if cell.value:
                    # Find the longest line in the cell after splitting by newline
                    lines = str(cell.value).split("\n")
                    max_length = max(max_length, max(len(line) for line in lines))
            except Exception as e:
                logger.exception("Error measuring cell length: %s", e)
        
        # Add a little buffer, but cap the width at a reasonable maximum (e.g., 60)
        adjusted_width = min((max_length + 2) * 1.2, 60)
        
        # Set a minimum width for short columns
        if adjusted_width < 15:
            adjusted_width = 15

        sheet.column_dimensions[column].width = adjusted_width

    # 8. Save
    excel_safe.save_workbook_safely(workbook, cloned_path)
    return cloned_path
